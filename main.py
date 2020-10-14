import pandas
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from collections import defaultdict, Counter

# Читаем файл ексль и результат передаем в переменную excel_data
# Переменная excel_data имеет тип <class 'pandas.core.frame.DataFrame'>
excel_data = pandas.read_excel('logs.xlsx', sheet_name='log')

# Преобразуем переменную excel_data в словарь с помощью метода to_dict()
# Результат передаем в переменную excel_data_dict
excel_data_dict = excel_data.to_dict(orient = 'records')

#print(excel_data_dict)
#создаем список товаров pr_list и заполняем его товарами из словаря excel_data_dict
pr_list = []
for i in range(len(excel_data_dict)):
    pr = excel_data_dict[i]['Купленные товары']
    pr_list.append(pr.split(','))
#Объединяем список списков в список new_list
new_list = [item for sublist in pr_list for item in sublist]
#print(new_list[:9])
product_counter = Counter(new_list) # Создаем объект Counter из полученного словаря и используем метод most_common
print(f'Самый популярный товар: {product_counter.most_common(7)}.')
prlist = product_counter.most_common(7)


sales_dict = defaultdict(int) # Создаем словарь с заранее заданным типом значений

for element in excel_data_dict:
    sales_dict[element['Браузер']] += 1

print(sales_dict)
sales_counter = Counter(sales_dict) # Создаем объект Counter из полученного словаря и используем метод most_common
print(f'Самый популярный браузер: {sales_counter.most_common(7)}.')
brlist = sales_counter.most_common(7)
print(brlist) 
print(brlist[0][1])

#Открыть книгу: 
wb = load_workbook(filename='report.xlsx')
#Активировать лист: 
sheet = wb['Лист1']
#Записать в ячейку:
sheet["A5"] = brlist[0][0]
sheet["A6"] = brlist[1][0]
sheet["A7"] = brlist[2][0]
sheet["A8"] = brlist[3][0]
sheet["A9"] = brlist[4][0]
sheet["A10"] = brlist[5][0]
sheet["A11"] = brlist[6][0]

sheet["B5"] = brlist[0][1]
sheet["B6"] = brlist[1][1]
sheet["B7"] = brlist[2][1]
sheet["B8"] = brlist[3][1]
sheet["B9"] = brlist[4][1]
sheet["B10"] = brlist[5][1]
sheet["B11"] = brlist[6][1]

sheet["A19"] = prlist[0][0]
sheet["A20"] = prlist[1][0]
sheet["A21"] = prlist[2][0]
sheet["A22"] = prlist[3][0]
sheet["A23"] = prlist[4][0]
sheet["A24"] = prlist[5][0]
sheet["A25"] = prlist[6][0]

sheet["B19"] = prlist[0][1]
sheet["B20"] = prlist[1][1]
sheet["B21"] = prlist[2][1]
sheet["B22"] = prlist[3][1]
sheet["B23"] = prlist[4][1]
sheet["B24"] = prlist[5][1]
sheet["B25"] = prlist[6][1]
#Сохранить книгу: 
wb.save(filename='report.xlsx')


#здесь была попытка вычислить количество посещений по месяцам:
#Пайтон заругался, что timestamp не имеет атрибута split =(
def get_value_from_list(object_list,month,brname):
    s=0
    t = pandas.tslib.Timestamp.now()
    for element in object_list:
        if month == t.day and element['Браузер'] == brname:
            s+=1
    # Возвращаем найденное значение
    return s
print(get_value_from_list(excel_data_dict,'01','Opera'))




